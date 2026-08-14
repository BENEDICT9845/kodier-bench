#!/usr/bin/env python3
"""Rebuild data/national.json from the InEK source files.

INPUTS (place in --raw, default: pipeline/raw/inek/):
  1. InEK DatenBrowser export  (.xlsx with a 'DRG-Verteilung' sheet)
       -> national Fälle per DRG, Nebendiagnosen (CC), Prozeduren (OPS), Krankenhausverteilung
       Source: https://datenbrowser.inek.org  (free account; export the default §21 dataset)
  2. InEK Fallpauschalen-Katalog (.xlsx, 'Hauptabteilungen' sheet)
       -> per-DRG cost weight (Bewertungsrelation) + mean LOS
       Source: https://www.g-drg.de  (aG-DRG-Katalog for the system year)

OUTPUT: data/national.json  (the national-reference layer the app embeds)

NOTE: raw InEK files are NOT committed (see .gitignore) — licence + size. Only the
derived aggregate (national.json) is tracked.

Usage:  python pipeline/build_national.py [--raw pipeline/raw/inek] [--out data/national.json]
"""
from __future__ import annotations
import argparse, glob, json, re
from pathlib import Path
import openpyxl

MDC = {"A":"Prä-MDC","B":"Nervensystem","C":"Auge","D":"HNO","E":"Atmungsorgane",
 "F":"Kreislauf / Herz","G":"Verdauungsorgane","H":"Leber, Galle, Pankreas","I":"Muskel-Skelett-System",
 "J":"Haut","K":"Endokrin & Stoffwechsel","L":"Niere & Harnwege","M":"Männl. Geschlechtsorgane",
 "N":"Weibl. Geschlechtsorgane","O":"Schwangerschaft & Geburt","P":"Neugeborene","Q":"Blut & blutbildende Organe",
 "R":"Neubildungen (hämat./solide)","S":"HIV","T":"Infektionen & Parasiten","U":"Psychische Krankheiten",
 "V":"Alkohol- & Drogengebrauch","W":"Polytrauma","X":"Verletzungen & Vergiftungen","Y":"Verbrennungen","Z":"Sonstige / Faktoren"}
DRG = re.compile(r"^[A-Z]\d{2}[A-Z]$")


def num(x):
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None


def find(raw: Path, needle_sheet: str):
    for f in raw.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            if needle_sheet in wb.sheetnames:
                return f
        except Exception:
            continue
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default="pipeline/raw/inek")
    ap.add_argument("--out", default="data/national.json")
    a = ap.parse_args()
    raw = Path(a.raw)

    export = find(raw, "DRG-Verteilung")
    katalog = find(raw, "Hauptabteilungen")
    if not export or not katalog:
        raise SystemExit(f"Missing inputs in {raw}: need a DatenBrowser export (DRG-Verteilung) "
                         f"and a Fallpauschalen-Katalog (Hauptabteilungen). See docstring.")

    wb = openpyxl.load_workbook(export, read_only=True, data_only=True)
    # case counts + LOS + label
    cases = {}
    for r in wb["DRG-Verteilung"].iter_rows(min_row=2, values_only=True):
        c = r[0]
        if isinstance(c, str) and DRG.match(c.strip()):
            c = c.strip()
            cases[c] = (int(num(r[2]) or 0), round(num(r[4]) or 0, 1),
                        re.sub(r"\s+", " ", str(r[1] or "")).strip()[:58])
    # cost weights
    wb2 = openpyxl.load_workbook(katalog, read_only=True, data_only=True)
    cw = {}
    for r in wb2["Hauptabteilungen"].iter_rows(min_row=7, values_only=True):
        c = r[0]
        if isinstance(c, str) and DRG.match(c.strip()):
            cw[c.strip()] = num(r[3])

    bases = {}
    for drg, (n, los, bez) in cases.items():
        b = drg[:3]
        bases.setdefault(b, {"mdc": drg[0], "label": bez, "tiers": []})
        bases[b]["tiers"].append({"drg": drg, "tier": drg[3], "cases": n, "los": los, "cw": cw.get(drg)})
    out = {}
    for b, v in bases.items():
        t = sorted(v["tiers"], key=lambda x: x["drg"])
        tot = sum(x["cases"] for x in t)
        for x in t:
            x["pct"] = round(100 * x["cases"] / tot, 1) if tot else 0
        cwc = sum(x["cases"] for x in t if x["cw"] is not None)
        mean_cw = round(sum(x["cases"] * x["cw"] for x in t if x["cw"] is not None) / cwc, 3) if cwc else None
        modal = max(t, key=lambda x: x["cases"])
        hi = None
        if modal["cw"] is not None and tot:
            hi = round(100 * sum(x["cases"] for x in t if x["cw"] is not None and x["cw"] > modal["cw"]) / tot, 1)
        out[b] = {"mdc": v["mdc"], "label": v["label"], "tiers": t, "total": tot,
                  "mean_cw": mean_cw, "hi_share": hi}
    allc = sum(x["cases"] for b in out for x in out[b]["tiers"] if x["cw"] is not None)
    nat_cmi = round(sum(x["cases"] * x["cw"] for b in out for x in out[b]["tiers"] if x["cw"] is not None) / allc, 3)

    def top(sheet, n=12):
        rows = []
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                rows.append({"code": str(r[0]), "label": str(r[1])[:52], "cases": int(num(r[2]) or 0)})
        return rows[:n]

    coh = [{"bed": str(r[0]), "own": str(r[1]), "pct": round((num(r[3]) or 0) * 100, 2)}
           for r in wb["Krankenhausverteilung"].iter_rows(min_row=2, values_only=True) if r and r[0]]

    total_cases = sum(x["cases"] for b in out for x in out[b]["tiers"])
    data = {"vintage": "§21 data Jan–Dec 2025 · aG-DRG 2026", "total_cases": total_cases,
            "nat_cmi": nat_cmi, "mdc": MDC, "bases": out, "cohort": coh,
            "top_cc": top("Nebendiagnosen"), "top_ops": top("Prozeduren")}
    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    json.dump(data, open(a.out, "w"), ensure_ascii=False, separators=(",", ":"))
    print(f"wrote {a.out}: {len(out)} DRG groups, national CMI {nat_cmi}")


if __name__ == "__main__":
    main()
